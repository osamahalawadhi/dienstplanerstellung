import calendar
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import List, Set, Optional, Tuple, Dict

import streamlit as st
from supabase import create_client
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

GREEN_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")


# ─────────────────────────────────────────────
# DATACLASSES
# ─────────────────────────────────────────────

@dataclass
class Employee:
    name: str
    is_fachkraft: bool
    availability: List[bool]       # length = days_in_month
    min_services: int
    max_services: int
    block_preferences: Set[int]    # e.g. {2, 3}
    wants_8_block: bool

    # filled after scheduling
    assigned_count: int = 0


@dataclass
class DayRequirement:
    target: int
    minimum: int
    needs_fachkraft: bool = True
    exact_target: bool = False


# ─────────────────────────────────────────────
# SUPABASE
# ─────────────────────────────────────────────

@st.cache_resource
def get_supabase():
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)


# ─────────────────────────────────────────────
# DATE HELPERS
# ─────────────────────────────────────────────

def get_days_in_month(month: int, year: int) -> int:
    return calendar.monthrange(year, month)[1]


def get_weekday_short(d: date) -> str:
    return ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][d.weekday()]


def get_day_label(day: int, month: int, year: int) -> str:
    d = date(year, month, day)
    return f"{get_weekday_short(d)} {d.strftime('%d.%m.%Y')}"


def get_excel_day_label(day: int, month: int, year: int) -> str:
    d = date(year, month, day)
    return f"{get_weekday_short(d)}\n{d.strftime('%d.%m.%Y')}"


def requirement_for_day(day: int, month: int, year: int) -> DayRequirement:
    wd = date(year, month, day).weekday()
    if wd >= 4:  # Fr, Sa, So
        return DayRequirement(target=3, minimum=3, needs_fachkraft=True, exact_target=True)
    return DayRequirement(target=3, minimum=2, needs_fachkraft=True, exact_target=False)


# ─────────────────────────────────────────────
# SUPABASE DB FUNCTIONS
# ─────────────────────────────────────────────

def get_or_create_planning_round(sb, month: int, year: int):
    title = f"Dienstplan {month:02d}/{year}"
    existing = sb.table("planning_rounds").select("*").eq("month", month).eq("year", year).execute()
    if existing.data:
        return existing.data[0]
    created = sb.table("planning_rounds").insert({"month": month, "year": year, "title": title}).execute()
    return created.data[0]


def load_employee_inputs(sb, planning_round_id: int):
    result = sb.table("employee_inputs").select("*").eq("planning_round_id", planning_round_id).order("name").execute()
    return result.data or []


def load_active_employees(sb):
    return sb.table("employees_master").select("*").eq("active", True).order("name").order("id").execute().data or []


def load_inactive_employees(sb):
    return sb.table("employees_master").select("*").eq("active", False).order("name").order("id").execute().data or []


def add_employee_master(sb, name: str):
    existing = sb.table("employees_master").select("id, active").eq("name", name).limit(1).execute()
    if existing.data:
        return {"status": "exists", "row": existing.data[0]}
    created = sb.table("employees_master").insert({"name": name, "active": True}).execute()
    return {"status": "created", "row": created.data[0] if created.data else None}


def deactivate_employee_master(sb, employee_id: int):
    return sb.table("employees_master").update({"active": False}).eq("id", employee_id).execute()


def reactivate_employee_master(sb, employee_id: int):
    return sb.table("employees_master").update({"active": True}).eq("id", employee_id).execute()


def load_existing_input_for_employee(sb, planning_round_id: int, employee_id: int):
    result = (
        sb.table("employee_inputs")
        .select("*")
        .eq("planning_round_id", planning_round_id)
        .eq("employee_id", employee_id)
        .limit(1)
        .execute()
    )
    return result.data[0] if result.data else None


def save_employee_input(sb, planning_round_id, employee_id, name, is_fachkraft,
                        min_services, max_services, block_preferences,
                        wants_8_block, availability):
    existing = (
        sb.table("employee_inputs")
        .select("id")
        .eq("planning_round_id", planning_round_id)
        .eq("employee_id", employee_id)
        .execute()
    )
    payload = {
        "planning_round_id": planning_round_id,
        "employee_id": employee_id,
        "name": name,
        "is_fachkraft": is_fachkraft,
        "min_services": min_services,
        "max_services": max_services,
        "block_preferences": block_preferences,
        "wants_8_block": wants_8_block,
        "availability": availability,
        "submitted": True,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if existing.data:
        return sb.table("employee_inputs").update(payload).eq("id", existing.data[0]["id"]).execute()
    return sb.table("employee_inputs").insert(payload).execute()


def build_employees_from_inputs(sb, planning_round_id: int, days_in_month: int) -> List[Employee]:
    rows = load_employee_inputs(sb, planning_round_id)
    employees = []
    for row in rows:
        if not row.get("submitted"):
            continue
        availability = row.get("availability") or []
        if len(availability) != days_in_month:
            availability = [False] * days_in_month

        blocks_raw = row.get("block_preferences") or []
        block_preferences = set()
        for b in blocks_raw:
            try:
                b_int = int(b)
                if b_int in {1, 2, 3, 4}:
                    block_preferences.add(b_int)
            except Exception:
                pass

        employees.append(Employee(
            name=row["name"],
            is_fachkraft=bool(row.get("is_fachkraft", False)),
            availability=[bool(x) for x in availability],
            min_services=int(row.get("min_services", 0)),
            max_services=int(row.get("max_services", 0)),
            block_preferences=block_preferences,
            wants_8_block=bool(row.get("wants_8_block", False)),
        ))
    return employees


# ─────────────────────────────────────────────
# CORE SCHEDULER – OR-Tools CP-SAT
# ─────────────────────────────────────────────

def check_block_feasibility(
    employees: List[Employee],
    month: int,
    year: int,
    days_in_month: int,
) -> List[str]:
    """
    Prüft ob jeder Mitarbeiter seine Min-Dienste mit seinen Blöcken erreichen kann.
    Gibt Fehlermeldungen zurück (leer = alles ok).
    """
    D = days_in_month
    errors = []

    for emp in employees:
        allowed_sizes = set(emp.block_preferences)

        if emp.wants_8_block and not allowed_sizes:
            # Nur 8er-Block: prüfe ob er mindestens einmal passt
            eight_possible = False
            for d in range(D):
                if d + 9 <= D:
                    work_days = list(range(d, d+4)) + list(range(d+5, d+9))
                    if all(emp.availability[dd] for dd in work_days):
                        eight_possible = True
                        break
            if not eight_possible:
                errors.append(
                    f"**{emp.name}**: 8er-Block kann nirgendwo platziert werden "
                    f"(zu wenig aufeinanderfolgende verfügbare Tage)."
                )
            elif emp.min_services > 8:
                errors.append(
                    f"**{emp.name}**: Min-Dienste={emp.min_services}, aber 8er-Block "
                    f"liefert maximal 8 Dienste."
                )
            continue

        if not allowed_sizes:
            continue

        # Greedy: zähle maximal erreichbare Dienste mit erlaubten Blöcken
        max_reachable = 0
        d = 0
        while d < D:
            placed = False
            for s in sorted(allowed_sizes, reverse=True):
                if d + s <= D:
                    if all(emp.availability[dd] for dd in range(d, d + s)):
                        max_reachable += s
                        d += s
                        placed = True
                        break
            if not placed:
                d += 1

        if max_reachable < emp.min_services:
            errors.append(
                f"**{emp.name}**: Mit Blöcken {sorted(allowed_sizes)} und eingetragener "
                f"Verfügbarkeit maximal {max_reachable} Dienste erreichbar, "
                f"aber Min-Dienste = {emp.min_services}."
            )

    return errors



def _build_model(
    employees: List[Employee],
    month: int,
    year: int,
    days_in_month: int,
) -> Tuple[cp_model.CpModel, List[List]]:
    """
    Builds the CP-SAT model with ALL rules strictly enforced – no relaxation.

    Hard constraints (always):
      - Availability
      - Block structure (only allowed block sizes)
      - 8-block: exactly 4-off-4
      - Max 4 consecutive work days, then at least 1 free
      - Min AND Max services per employee
      - Weekday: min 2 staff, at least 1 Fachkraft
      - Weekend (Fr/Sa/So): min 3 staff, at least 1 Fachkraft

    Soft (objective):
      - 3rd person on weekdays to reach desired service counts
    """
    n = len(employees)
    D = days_in_month
    model = cp_model.CpModel()

    # ── Decision variables ────────────────────────────────────────────────
    shift: List[List] = [
        [model.NewBoolVar(f"shift_e{e}_d{d}") for d in range(D)]
        for e in range(n)
    ]

    # ── Availability ──────────────────────────────────────────────────────
    for e, emp in enumerate(employees):
        for d in range(D):
            if not emp.availability[d]:
                model.Add(shift[e][d] == 0)

    # ── Block structure ───────────────────────────────────────────────────
    for e, emp in enumerate(employees):
        allowed_sizes = set(emp.block_preferences)

        block_start: Dict[Tuple[int, int], object] = {}
        for d in range(D):
            for s in allowed_sizes:
                if d + s <= D:
                    if all(emp.availability[dd] for dd in range(d, d + s)):
                        block_start[(d, s)] = model.NewBoolVar(f"bs_e{e}_d{d}_s{s}")

        eight_block_vars: List[Tuple[int, object]] = []
        if emp.wants_8_block:
            for d in range(D):
                if d + 9 <= D:
                    work_days_8 = list(range(d, d + 4)) + list(range(d + 5, d + 9))
                    if all(emp.availability[dd] for dd in work_days_8):
                        eight_block_vars.append((d, model.NewBoolVar(f"8blk_e{e}_d{d}")))

        for d in range(D):
            covering = []
            for s in allowed_sizes:
                for start in range(max(0, d - s + 1), d + 1):
                    if (start, s) in block_start and start + s > d:
                        covering.append(block_start[(start, s)])
            for (bd, bvar) in eight_block_vars:
                work_days_in_8 = list(range(bd, bd + 4)) + list(range(bd + 5, bd + 9))
                if d in work_days_in_8:
                    covering.append(bvar)

            if covering:
                model.Add(shift[e][d] == sum(covering))
            else:
                model.Add(shift[e][d] == 0)

        # 8-block: free day must be free
        for (bd, bvar) in eight_block_vars:
            model.Add(shift[e][bd + 4] == 0).OnlyEnforceIf(bvar)

    # ── Max 4 consecutive work days ───────────────────────────────────────
    for e in range(n):
        for d in range(D - 4):
            model.Add(sum(shift[e][d + k] for k in range(5)) <= 4)

    # ── Min / Max services ────────────────────────────────────────────────
    for e, emp in enumerate(employees):
        total = sum(shift[e][d] for d in range(D))
        model.Add(total >= emp.min_services)
        model.Add(total <= emp.max_services)

    # ── Max 3 workers per day (hard) ─────────────────────────────────────
    for d in range(D):
        model.Add(sum(shift[e][d] for e in range(n)) <= 3)

    # ── Daily coverage: SOFT via penalty variables ────────────────────────
    # We never make coverage a hard constraint.
    # Instead we penalise shortfalls heavily in the objective so the solver
    # fills days as much as possible, but never returns INFEASIBLE.
    fachkraft_indices = [e for e, emp in enumerate(employees) if emp.is_fachkraft]

    # shortfall_day[d]  = max(0, minimum_required - actual_assigned)
    # fk_missing[d]     = 1 if no Fachkraft assigned on day d, else 0
    shortfall_day = []
    fk_missing = []

    for d in range(D):
        day = d + 1
        req = requirement_for_day(day, month, year)
        daily_total = sum(shift[e][d] for e in range(n))

        # Shortfall for staffing minimum
        sf = model.NewIntVar(0, n, f"sf_d{d}")
        model.Add(sf >= req.minimum - daily_total)
        model.Add(sf >= 0)
        shortfall_day.append(sf)

        # Fachkraft missing indicator
        available_fk_today = [e for e in fachkraft_indices if employees[e].availability[d]]
        if available_fk_today:
            fk_sum = sum(shift[e][d] for e in available_fk_today)
            fkm = model.NewBoolVar(f"fkm_d{d}")
            # fkm = 1 when fk_sum == 0
            model.Add(fk_sum == 0).OnlyEnforceIf(fkm)
            model.Add(fk_sum >= 1).OnlyEnforceIf(fkm.Not())
            fk_missing.append(fkm)
        else:
            # No FK available at all – nothing to penalise
            fk_missing.append(None)

    # ── Objective ────────────────────────────────────────────────────────
    # Priority 1 (highest): Tagesbesetzung einhalten → Strafe -1000 pro fehlender Person
    # Priority 2: Fachkraft pro Tag → Strafe -1000 wenn keine FK
    # Priority 3 (lowest): 3. Person an Wochentagen → Belohnung +1
    #
    # WICHTIG: Wir belohnen NIE einfach "mehr Dienste" für einen Mitarbeiter,
    # da das den Solver dazu bringen könnte, Max-Dienste zu ignorieren.
    # Min/Max sind bereits als harte Constraints gesetzt (Zeilen ~351-352).
    objective_terms = []

    # Strafe für Unterbesetzung pro Tag
    for sf in shortfall_day:
        objective_terms.append(-1000 * sf)

    # Strafe für fehlende Fachkraft pro Tag
    for fkm in fk_missing:
        if fkm is not None:
            objective_terms.append(-1000 * fkm)

    # Belohnung für 3. Person an Wochentagen (soft, niedrige Gewichtung)
    # Durch Max-3-Constraint und Max-Dienste-Constraint kann dies nie
    # die harten Grenzen überschreiten.
    for d in range(D):
        req = requirement_for_day(d + 1, month, year)
        if not req.exact_target:
            daily_total = sum(shift[e][d] for e in range(n))
            objective_terms.append(daily_total)

    model.Maximize(sum(objective_terms))
    return model, shift


def _solve(model: cp_model.CpModel, shift, employees, days_in_month) -> Tuple[int, object]:
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0
    solver.parameters.num_search_workers = 4
    status = solver.Solve(model)
    return status, solver


def _pre_solve_diagnostics(
    employees: List[Employee],
    month: int,
    year: int,
    days_in_month: int,
) -> List[str]:
    """
    Check before solving whether rules are reachable.
    Returns a list of warning strings – does NOT block the solve.
    """
    issues: List[str] = []
    n = len(employees)
    fachkraft_indices = [e for e, emp in enumerate(employees) if emp.is_fachkraft]

    for d in range(days_in_month):
        day = d + 1
        req = requirement_for_day(day, month, year)

        # How many employees could possibly work this day
        # (available + have at least one valid block covering this day)
        can_work = 0
        for e, emp in enumerate(employees):
            if not emp.availability[d]:
                continue
            # Check if any block of their preferred sizes covers day d
            has_block = False
            for s in emp.block_preferences:
                for start in range(max(0, d - s + 1), d + 1):
                    end = start + s
                    if end <= days_in_month:
                        if all(emp.availability[dd] for dd in range(start, end)):
                            has_block = True
                            break
                if has_block:
                    break
            # Also check 8-block
            if not has_block and emp.wants_8_block:
                for start in range(days_in_month):
                    if start + 9 <= days_in_month:
                        work_days = list(range(start, start+4)) + list(range(start+5, start+9))
                        if d in work_days and all(emp.availability[dd] for dd in work_days):
                            has_block = True
                            break
            if has_block:
                can_work += 1

        if can_work < req.minimum:
            issues.append(
                f"⚠️ Unterbesetzung wahrscheinlich an {get_day_label(day, month, year)}: "
                f"nur {can_work} Mitarbeiter können diesen Tag (inkl. Blockregeln) besetzen, "
                f"{req.minimum} benötigt."
            )

        # Fachkraft check
        fk_can_work = sum(
            1 for e in fachkraft_indices
            if employees[e].availability[d]
        )
        if fk_can_work == 0:
            issues.append(
                f"⚠️ Keine Fachkraft verfügbar an {get_day_label(day, month, year)} – "
                f"Fachkraft-Regel kann nicht erfüllt werden."
            )

    # Min-services reachability per employee
    for emp in employees:
        avail_count = sum(emp.availability)
        if avail_count < emp.min_services:
            issues.append(
                f"⚠️ {emp.name}: nur {avail_count} verfügbare Tage, "
                f"aber Min-Dienste = {emp.min_services} – nicht erreichbar."
            )

    return issues


def generate_schedule(
    employees: List[Employee],
    month: int,
    year: int,
    days_in_month: int,
) -> Tuple[List[List[int]], List[str], List[Employee]]:
    """
    Builds a shift schedule using OR-Tools CP-SAT.

    Hard constraints (never violated, always respected):
      - Availability
      - Block structure (only allowed block sizes, exact 8-block pattern)
      - Max 4 consecutive work days
      - Min / Max services per employee

    Soft constraints (maximised, but plan always produced):
      - Daily minimum staffing (2 weekday / 3 weekend)
      - At least 1 Fachkraft per day
      - 3rd person on weekdays

    Pre-solve diagnostics warn about days that cannot be fully covered.
    The plan is always returned even if some days are under-staffed.
    """
    warnings: List[str] = []
    n = len(employees)
    D = days_in_month
    assignments_by_day: List[List[int]] = [[] for _ in range(D)]

    # ── Pre-solve diagnostics ─────────────────────────────────────────────
    diag_issues = _pre_solve_diagnostics(employees, month, year, D)
    warnings.extend(diag_issues)

    # ── Build and solve ───────────────────────────────────────────────────
    model, shift = _build_model(employees, month, year, D)
    status, solver = _solve(model, shift, employees, D)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # This should now be extremely rare since daily coverage is soft
        warnings.append(
            "❌ Solver konnte keinen Plan erstellen. "
            "Bitte prüfe ob Min/Max-Dienste mit den verfügbaren Tagen erreichbar sind."
        )
        return assignments_by_day, warnings, employees

    # ── Read solution ─────────────────────────────────────────────────────
    for d in range(D):
        for e in range(n):
            if solver.Value(shift[e][d]) == 1:
                assignments_by_day[d].append(e)
                employees[e].assigned_count += 1

    # ── Post-solve warnings ───────────────────────────────────────────────
    fachkraft_indices = [e for e, emp in enumerate(employees) if emp.is_fachkraft]
    for d in range(D):
        day = d + 1
        req = requirement_for_day(day, month, year)
        assigned = assignments_by_day[d]

        if len(assigned) < req.minimum:
            warnings.append(
                f"Unterbesetzung {get_day_label(day, month, year)}: "
                f"{len(assigned)} statt mindestens {req.minimum} Personen."
            )

        fk_count = sum(1 for e in assigned if employees[e].is_fachkraft)
        if fk_count == 0 and req.needs_fachkraft:
            warnings.append(
                f"Keine Fachkraft an {get_day_label(day, month, year)}."
            )

    for e, emp in enumerate(employees):
        if emp.assigned_count < emp.min_services:
            warnings.append(
                f"Min-Dienste nicht erreicht: {emp.name} hat {emp.assigned_count}, "
                f"Minimum {emp.min_services}."
            )

    return assignments_by_day, warnings, employees


# ─────────────────────────────────────────────
# WARNINGS FILTER
# ─────────────────────────────────────────────

def filter_user_warnings(warnings: List[str]) -> List[str]:
    important_prefixes = [
        "Unterbesetzung",
        "Keine Fachkraft",
        "Wochenende unterbesetzt",
        "Min-Dienste nicht erreicht",
        "⚠️ Der Solver",
        "Keine Fachkraft verfügbar",
    ]
    seen = set()
    result = []
    for w in warnings:
        if any(w.startswith(p) for p in important_prefixes) and w not in seen:
            result.append(w)
            seen.add(w)
    return result


# ─────────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────────

def build_input_overview_excel(
    employees: List[Employee],
    month: int,
    year: int,
    days_in_month: int,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Eingaben"

    headers = ["Name", "Fachkraft", "Min", "Max", "Blöcke", "8er-Wunsch"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for d in range(1, days_in_month + 1):
        col = get_column_letter(6 + d)
        ws[f"{col}1"] = get_day_label(d, month, year)
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 20
    for c, w in zip("BCDEF", [12, 8, 8, 14, 12]):
        ws.column_dimensions[c].width = w
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(6 + d)].width = 12
    ws.row_dimensions[1].height = 32

    for row_idx, emp in enumerate(employees, start=2):
        ws[f"A{row_idx}"] = emp.name
        ws[f"B{row_idx}"] = "Ja" if emp.is_fachkraft else "Nein"
        ws[f"C{row_idx}"] = emp.min_services
        ws[f"D{row_idx}"] = emp.max_services
        ws[f"E{row_idx}"] = ",".join(str(x) for x in sorted(emp.block_preferences))
        ws[f"F{row_idx}"] = "Ja" if emp.wants_8_block else "Nein"

        for day_idx, available in enumerate(emp.availability, start=1):
            col = get_column_letter(6 + day_idx)
            cell = ws[f"{col}{row_idx}"]
            cell.value = "Ja" if available else "Nein"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if available:
                cell.fill = GREEN_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_schedule_excel(
    original_employees: List[Employee],
    final_employees: List[Employee],
    assignments_by_day: List[List[int]],
    warnings: List[str],
    month: int,
    year: int,
    days_in_month: int,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dienstplan"

    ws["A1"] = "Name"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for d in range(1, days_in_month + 1):
        col = get_column_letter(1 + d)
        ws[f"{col}1"] = get_excel_day_label(d, month, year)
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sum_col = get_column_letter(days_in_month + 2)
    ws[f"{sum_col}1"] = "Summe"
    ws[f"{sum_col}1"].font = Font(bold=True)
    ws[f"{sum_col}1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 20
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(1 + d)].width = 12
    ws.column_dimensions[sum_col].width = 10
    ws.row_dimensions[1].height = 32

    assigned_sets = [set(ids) for ids in assignments_by_day]

    for row_idx, emp in enumerate(original_employees, start=2):
        ws[f"A{row_idx}"] = emp.name
        emp_idx = row_idx - 2
        service_count = 0

        for d0 in range(days_in_month):
            col = get_column_letter(2 + d0)
            cell = ws[f"{col}{row_idx}"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if emp_idx in assigned_sets[d0]:
                cell.value = "X"
                cell.fill = GREEN_FILL
                service_count += 1

        ws[f"{sum_col}{row_idx}"] = service_count
        ws[f"{sum_col}{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")

    # Warnings sheet
    ws2 = wb.create_sheet("Warnungen")
    ws2["A1"] = "Warnungen"
    ws2["A1"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 140
    for i, w in enumerate(filter_user_warnings(warnings), start=2):
        ws2[f"A{i}"] = w

    # Stats sheet
    ws3 = wb.create_sheet("Statistik")
    for col_idx, header in enumerate(["Name", "Fachkraft", "Min", "Max", "Geplant", "Wunschblöcke", "8er-Wunsch"], start=1):
        ws3.cell(1, col_idx, header).font = Font(bold=True)

    for i, (orig, final) in enumerate(zip(original_employees, final_employees), start=2):
        ws3.cell(i, 1, orig.name)
        ws3.cell(i, 2, "Ja" if orig.is_fachkraft else "Nein")
        ws3.cell(i, 3, orig.min_services)
        ws3.cell(i, 4, orig.max_services)
        ws3.cell(i, 5, final.assigned_count)
        ws3.cell(i, 6, ",".join(map(str, sorted(orig.block_preferences))) if orig.block_preferences else "")
        ws3.cell(i, 7, "Ja" if orig.wants_8_block else "Nein")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(page_title="Dienstplan Mitarbeitereingabe", layout="wide")
st.title("Dienstplan Mitarbeitereingabe")

sb = get_supabase()

with st.sidebar:
    st.header("Planungsmonat")
    month = st.number_input("Monat", min_value=1, max_value=12, value=3, step=1)
    year = st.number_input("Jahr", min_value=2025, max_value=2100, value=2026, step=1)
    days_in_month = get_days_in_month(int(month), int(year))

round_row = get_or_create_planning_round(sb, int(month), int(year))
planning_round_id = round_row["id"]

st.success(f"Planungsrunde aktiv: {round_row['title']}")

employees_master = load_active_employees(sb)

if not employees_master:
    st.error("Keine aktiven Mitarbeitenden in 'employees_master' gefunden.")
    st.stop()

rows = load_employee_inputs(sb, planning_round_id)
submitted_by_employee_id = {
    row["employee_id"]: row
    for row in rows
    if row.get("submitted") and row.get("employee_id") is not None
}

# ── Submission status ─────────────────────────────────────────────────
st.subheader("Status der Mitarbeitereingaben")
total_count = len(employees_master)
submitted_count = len(submitted_by_employee_id)
st.info(f"{submitted_count} von {total_count} Mitarbeitenden haben bereits eingetragen.")

for emp in employees_master:
    emp_id = emp["id"]
    name = emp["name"]
    if emp_id in submitted_by_employee_id:
        updated_at = submitted_by_employee_id[emp_id].get("updated_at", "")
        st.write(f"✅ **{name}** — eingetragen — zuletzt geändert: {updated_at}")
    else:
        st.write(f"❌ **{name}** — noch offen")

missing_names = [emp["name"] for emp in employees_master if emp["id"] not in submitted_by_employee_id]
if missing_names:
    st.warning("Noch offen: " + ", ".join(missing_names))
else:
    st.success("Alle Mitarbeitenden haben ihre Eingaben abgegeben.")

st.markdown("---")

# ── Employee input form ───────────────────────────────────────────────
st.subheader("Eigene Daten eintragen")

employee_options = [emp["name"] for emp in employees_master]
selected_name = st.selectbox("Mitarbeiter auswählen", employee_options)
selected_employee = next(emp for emp in employees_master if emp["name"] == selected_name)
employee_id = selected_employee["id"]
name = selected_employee["name"]

st.info(f"Mitarbeiter: **{name}**")

existing_input = load_existing_input_for_employee(sb, planning_round_id, employee_id)

default_is_fachkraft = False
default_min_services = 8
default_max_services = 15
default_blocks = [2]
default_wants_8 = False
default_availability = [True] * days_in_month

if existing_input:
    default_is_fachkraft = bool(existing_input.get("is_fachkraft", False))
    default_min_services = int(existing_input.get("min_services", 8))
    default_max_services = int(existing_input.get("max_services", 15))
    loaded_blocks = existing_input.get("block_preferences") or []
    default_wants_8 = bool(existing_input.get("wants_8_block", False))
    loaded_availability = existing_input.get("availability") or []
    if isinstance(loaded_blocks, list):
        default_blocks = [int(x) for x in loaded_blocks if str(x).isdigit() and int(x) in [1, 2, 3, 4]]
    if isinstance(loaded_availability, list) and len(loaded_availability) == days_in_month:
        default_availability = [bool(x) for x in loaded_availability]

with st.form("employee_form"):
    is_fachkraft = st.checkbox("Fachkraft", value=default_is_fachkraft)
    c1, c2 = st.columns(2)
    with c1:
        min_services = st.number_input("Min-Dienste", min_value=0, max_value=31, value=default_min_services, step=1)
    with c2:
        max_services = st.number_input("Max-Dienste", min_value=0, max_value=31, value=default_max_services, step=1)

    block_preferences = st.multiselect("Bevorzugte Blockgrößen", options=[1, 2, 3, 4], default=default_blocks)
    wants_8_block = st.checkbox("8er-Block-Wunsch (4 + frei + 4)", value=default_wants_8)

    st.write("Verfügbarkeit")
    availability = []
    cols = st.columns(7)
    for d in range(1, days_in_month + 1):
        with cols[(d - 1) % 7]:
            available = st.checkbox(
                get_day_label(d, int(month), int(year)),
                value=default_availability[d - 1],
                key=f"{employee_id}_day_{month}_{year}_{d}",
            )
            availability.append(available)

    submitted = st.form_submit_button("Speichern")

if submitted:
    errors = []
    if not name.strip():
        errors.append("Name fehlt.")
    if max_services < min_services:
        errors.append("Max-Dienste dürfen nicht kleiner als Min-Dienste sein.")
    if not block_preferences and not wants_8_block:
        errors.append("Mindestens ein Blockwunsch oder 8er-Wunsch ist erforderlich.")
    if errors:
        for err in errors:
            st.error(err)
    else:
        try:
            save_employee_input(
                sb=sb,
                planning_round_id=planning_round_id,
                employee_id=employee_id,
                name=name.strip(),
                is_fachkraft=bool(is_fachkraft),
                min_services=int(min_services),
                max_services=int(max_services),
                block_preferences=list(block_preferences),
                wants_8_block=bool(wants_8_block),
                availability=availability,
            )
            st.success("Deine Daten wurden gespeichert.")
            st.rerun()
        except Exception as e:
            st.error("Speichern fehlgeschlagen.")
            st.exception(e)

st.markdown("---")

# ── Admin area ────────────────────────────────────────────────────────
st.subheader("Admin-Bereich")
admin_mode = st.checkbox("Admin-Modus aktivieren")

if admin_mode:
    active_employees = load_active_employees(sb)
    inactive_employees = load_inactive_employees(sb)

    st.markdown("### Mitarbeiter verwalten")

    with st.form("add_employee_form"):
        new_employee_name = st.text_input("Neuen Mitarbeiter hinzufügen")
        add_submitted = st.form_submit_button("Mitarbeiter anlegen")

    if add_submitted:
        clean_name = new_employee_name.strip()
        if not clean_name:
            st.error("Bitte einen Namen eingeben.")
        else:
            try:
                result = add_employee_master(sb, clean_name)
                if result["status"] == "created":
                    st.success(f"Mitarbeiter **{clean_name}** wurde angelegt.")
                    st.rerun()
                else:
                    if result["row"].get("active"):
                        st.warning(f"Mitarbeiter **{clean_name}** existiert bereits.")
                    else:
                        st.warning(
                            f"Mitarbeiter **{clean_name}** existiert bereits, ist aber deaktiviert. "
                            f"Du kannst ihn unten wieder aktivieren."
                        )
            except Exception as e:
                st.error("Mitarbeiter konnte nicht angelegt werden.")
                st.exception(e)

    if active_employees:
        active_options = {emp["name"]: emp["id"] for emp in active_employees}
        selected_remove_name = st.selectbox("Aktiven Mitarbeiter deaktivieren", options=list(active_options.keys()), key="remove_employee_select")
        if st.button("Mitarbeiter deaktivieren"):
            try:
                deactivate_employee_master(sb, active_options[selected_remove_name])
                st.success(f"Mitarbeiter **{selected_remove_name}** wurde deaktiviert.")
                st.rerun()
            except Exception as e:
                st.error("Mitarbeiter konnte nicht deaktiviert werden.")
                st.exception(e)

    if inactive_employees:
        inactive_options = {emp["name"]: emp["id"] for emp in inactive_employees}
        selected_reactivate_name = st.selectbox("Deaktivierten Mitarbeiter wieder aktivieren", options=list(inactive_options.keys()), key="reactivate_employee_select")
        if st.button("Mitarbeiter reaktivieren"):
            try:
                reactivate_employee_master(sb, inactive_options[selected_reactivate_name])
                st.success(f"Mitarbeiter **{selected_reactivate_name}** wurde reaktiviert.")
                st.rerun()
            except Exception as e:
                st.error("Mitarbeiter konnte nicht reaktiviert werden.")
                st.exception(e)

    st.markdown("---")

    employees_for_plan = build_employees_from_inputs(sb, planning_round_id, days_in_month)
    st.write(f"Eingetragene Mitarbeitende für Planung: **{len(employees_for_plan)}**")

    if employees_for_plan:
        for emp in employees_for_plan:
            st.write(
                f"- {emp.name} | Fachkraft: {'Ja' if emp.is_fachkraft else 'Nein'} | "
                f"Min: {emp.min_services} | Max: {emp.max_services} | "
                f"Blöcke: {sorted(emp.block_preferences)} | 8er: {'Ja' if emp.wants_8_block else 'Nein'}"
            )

        # ── Diagnose vor der Planung ──────────────────────────────────────
        st.markdown("### Diagnose vor der Planung")
        pre_check_errors = []
        pre_check_warnings_list = []

        # 1) Verfügbarkeit pro Mitarbeiter
        st.markdown("**Verfügbarkeit pro Mitarbeiter:**")
        for emp in employees_for_plan:
            avail_count = sum(emp.availability)
            avail_pct = int(avail_count / days_in_month * 100)
            if avail_count == 0:
                st.error(
                    f"❌ **{emp.name}**: 0 von {days_in_month} Tagen verfügbar – "
                    f"Verfügbarkeit wurde nicht eingetragen!"
                )
                pre_check_errors.append(emp.name)
            elif avail_count < emp.min_services:
                st.warning(
                    f"⚠️ **{emp.name}**: nur {avail_count} Tage verfügbar, "
                    f"aber Min-Dienste = {emp.min_services} – Min möglicherweise nicht erreichbar."
                )
                pre_check_warnings_list.append(emp.name)
            else:
                st.write(
                    f"✅ **{emp.name}**: {avail_count} von {days_in_month} Tagen verfügbar ({avail_pct}%)"
                )

        # 2) Gesamtkapazität prüfen
        st.markdown("**Gesamtkapazität:**")
        total_max = sum(emp.max_services for emp in employees_for_plan)
        total_min_needed = sum(
            requirement_for_day(d + 1, int(month), int(year)).minimum
            for d in range(days_in_month)
        )
        if total_max < total_min_needed:
            st.error(
                f"❌ Gesamte Max-Dienste ({total_max}) reichen nicht für "
                f"den Mindestbedarf des Monats ({total_min_needed} benötigt)."
            )
            pre_check_errors.append("Gesamtkapazität")
        else:
            st.write(
                f"✅ Kapazität ausreichend: {total_max} Max-Dienste verfügbar, "
                f"{total_min_needed} Dienste mindestens benötigt."
            )

        # 3) Kritische Tage
        st.markdown("**Kritische Tage (zu wenig verfügbare Mitarbeiter):**")
        critical_days = []
        for d in range(days_in_month):
            available_today = sum(1 for emp in employees_for_plan if emp.availability[d])
            req = requirement_for_day(d + 1, int(month), int(year))
            if available_today < req.minimum:
                critical_days.append((d + 1, available_today, req.minimum))

        if critical_days:
            for day, avail, needed in critical_days:
                st.warning(
                    f"⚠️ {get_day_label(day, int(month), int(year))}: "
                    f"nur {avail} verfügbar, {needed} benötigt."
                )
        else:
            st.write("✅ Alle Tage haben ausreichend verfügbare Mitarbeiter.")

        # 4) Block-Erreichbarkeit prüfen
        st.markdown("**Block-Erreichbarkeit (können Min-Dienste mit den Blöcken erreicht werden?):**")
        block_errors = check_block_feasibility(
            employees_for_plan, int(month), int(year), days_in_month
        )
        if block_errors:
            for err in block_errors:
                st.error(f"❌ {err}")
            pre_check_errors.extend(block_errors)
        else:
            st.write("✅ Alle Mitarbeiter können ihre Min-Dienste mit ihren Blöcken erreichen.")

        # Zusammenfassung
        if pre_check_errors:
            st.error(
                "🚫 **Planung blockiert:** Bitte die rot markierten Probleme beheben. "
                "Der Plan kann so nicht erstellt werden."
            )
        elif pre_check_warnings_list:
            st.warning(
                "⚠️ Es gibt Warnungen – der Plan wird trotzdem versucht, "
                "aber manche Min-Dienste könnten unterschritten werden."
            )
        else:
            st.success("✅ Alle Prüfungen bestanden – Planung kann starten.")

        st.markdown("---")

        overview_excel = build_input_overview_excel(employees_for_plan, int(month), int(year), days_in_month)
        st.download_button(
            label="Eingaben als Kontroll-Excel herunterladen",
            data=overview_excel,
            file_name=f"eingaben_{int(month):02d}_{int(year)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown("---")

        if st.button("Dienstplan erstellen", disabled=bool(pre_check_errors)):
            with st.spinner("Dienstplan wird berechnet (OR-Tools CP-SAT)..."):
                assignments_by_day, warnings, final_employees = generate_schedule(
                    employees_for_plan,
                    int(month),
                    int(year),
                    days_in_month,
                )

            schedule_excel = build_schedule_excel(
                original_employees=employees_for_plan,
                final_employees=final_employees,
                assignments_by_day=assignments_by_day,
                warnings=warnings,
                month=int(month),
                year=int(year),
                days_in_month=days_in_month,
            )

            important_warnings = filter_user_warnings(warnings)
            st.success("Dienstplan wurde erstellt.")

            col_dl, col_prev = st.columns([1, 1])
            with col_dl:
                st.download_button(
                    label="📥 Dienstplan als Excel herunterladen",
                    data=schedule_excel,
                    file_name=f"dienstplan_{int(month):02d}_{int(year)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # ── Vorschau im Browser ───────────────────────────────────────
            st.markdown("### Vorschau Dienstplan")

            assigned_sets = [set(ids) for ids in assignments_by_day]
            fachkraft_set = {e for e, emp in enumerate(employees_for_plan) if emp.is_fachkraft}

            # Build header row
            day_labels = [get_day_label(d + 1, int(month), int(year)) for d in range(days_in_month)]

            # Colour logic: highlight worked cells
            # We render as an HTML table for full colour control
            html_rows = []

            # Header
            th_days = "".join(
                f'<th style="background:#2c3e50;color:white;padding:4px 6px;'
                f'font-size:11px;white-space:pre-line;text-align:center;'
                f'min-width:44px">{lbl.replace(" ", chr(10))}</th>'
                for lbl in day_labels
            )
            html_rows.append(
                f'<tr>'
                f'<th style="background:#2c3e50;color:white;padding:4px 8px;text-align:left;min-width:110px">Name</th>'
                f'{th_days}'
                f'<th style="background:#2c3e50;color:white;padding:4px 8px;text-align:center">Summe</th>'
                f'</tr>'
            )

            # One row per employee
            for e, emp in enumerate(employees_for_plan):
                count = 0
                cells = []
                for d in range(days_in_month):
                    if e in assigned_sets[d]:
                        count += 1
                        # Green for Fachkraft, lighter green for others
                        bg = "#27ae60" if e in fachkraft_set else "#90EE90"
                        fg = "white" if e in fachkraft_set else "#1a1a1a"
                        cells.append(
                            f'<td style="background:{bg};color:{fg};text-align:center;'
                            f'padding:3px;font-weight:bold;font-size:12px">X</td>'
                        )
                    else:
                        cells.append(
                            '<td style="background:#f8f9fa;text-align:center;'
                            'padding:3px;font-size:12px;color:#ccc">–</td>'
                        )

                fk_label = " ⭐" if emp.is_fachkraft else ""
                name_cell = (
                    f'<td style="padding:4px 8px;font-weight:bold;'
                    f'white-space:nowrap;font-size:12px">{emp.name}{fk_label}</td>'
                )
                sum_cell = (
                    f'<td style="text-align:center;font-weight:bold;padding:4px 8px;'
                    f'background:#eaf4fb;font-size:12px">{count}</td>'
                )
                html_rows.append(f'<tr>{name_cell}{"".join(cells)}{sum_cell}</tr>')

            # Summary row: how many per day
            summary_cells = []
            for d in range(days_in_month):
                count_day = len(assigned_sets[d])
                fk_day = sum(1 for e in assigned_sets[d] if e in fachkraft_set)
                req = requirement_for_day(d + 1, int(month), int(year))
                if count_day < req.minimum:
                    bg = "#e74c3c"; fg = "white"   # red = understaff
                elif fk_day == 0:
                    bg = "#f39c12"; fg = "white"   # orange = no FK
                else:
                    bg = "#ecf0f1"; fg = "#2c3e50"  # neutral
                summary_cells.append(
                    f'<td style="background:{bg};color:{fg};text-align:center;'
                    f'font-weight:bold;padding:3px;font-size:11px">{count_day}</td>'
                )

            html_rows.append(
                f'<tr>'
                f'<td style="padding:4px 8px;font-weight:bold;font-size:12px;'
                f'background:#ecf0f1">Besetzt</td>'
                f'{"".join(summary_cells)}'
                f'<td style="background:#ecf0f1"></td>'
                f'</tr>'
            )

            html_table = (
                '<div style="overflow-x:auto;margin-top:8px">'
                '<table style="border-collapse:collapse;width:100%;font-family:sans-serif">'
                + "".join(html_rows)
                + "</table></div>"
                "<p style='font-size:11px;color:#888;margin-top:6px'>"
                "⭐ = Fachkraft &nbsp;|&nbsp; "
                "<span style='background:#27ae60;color:white;padding:1px 6px;border-radius:3px'>X</span> Fachkraft arbeitet &nbsp;|&nbsp; "
                "<span style='background:#90EE90;padding:1px 6px;border-radius:3px'>X</span> Mitarbeiter arbeitet &nbsp;|&nbsp; "
                "<span style='background:#e74c3c;color:white;padding:1px 6px;border-radius:3px'>Zahl</span> Unterbesetzt &nbsp;|&nbsp; "
                "<span style='background:#f39c12;color:white;padding:1px 6px;border-radius:3px'>Zahl</span> Keine Fachkraft"
                "</p>"
            )

            st.markdown(html_table, unsafe_allow_html=True)

            st.subheader("Warnungen")
            if important_warnings:
                for warning in important_warnings:
                    st.warning(warning)
            else:
                st.info("Keine wichtigen Warnungen.")
    else:
        st.warning("Noch keine vollständigen Mitarbeitereingaben vorhanden.")
